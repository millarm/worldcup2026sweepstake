"""Generate ``wcsweepstake/tournament_data.json`` from the source spreadsheet.

The website's logic is a faithful re-implementation of the Excel workbook
``data/World_Cup_sweepstake.xlsx``.  Rather than hand-copy 48 teams, 72 group
fixtures and the knockout bracket, we extract them straight from the workbook so
the data can be regenerated if the sweepstake spreadsheet changes.

Run with::

    python scripts/generate_data.py

This requires ``openpyxl`` (a dev-only dependency, see requirements-dev.txt).
The generated JSON is committed to the repo so the runtime has no dependency on
openpyxl or the spreadsheet itself.
"""
from __future__ import annotations

import datetime
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "World_Cup_sweepstake.xlsx"
OUT = ROOT / "wcsweepstake" / "tournament_data.json"

# Short owner names used in the Tracker sheet -> full names from the Teams sheet.
OWNER_FULL_NAMES = {
    "Tish": "Atish Nazir",
    "Ben": "Ben Eveson",
    "Caoimhe": "Caoimhe Whelan",
    "Div": "Div Saxena",
    "Fabio": "Fabio Allegra",
    "Jane": "Jane Stinson",
    "Katia": "Katia Newman",
    "Lucy": "Lucy Philpott",
    "Marija": "Marija Skramic",
    "Mark": "Mark Evans",
    "Matt M": "Matt Millar",
    "Matt P": "Matt Perry",
    "Mike": "Mike Jackson",
    "Nelly": "Nelly Trakidou",
    "Nick": "Nick Robinson",
    "Rich": "Richard Ehlen",
    "Scott": "Scott Wallman",
    "Sean": "Sean Gill",
    "Tina": "Tina Louis",
    "Tom": "Tom Low",
}

# Teams-sheet spellings -> canonical Tracker spellings.
TEAM_ALIASES = {
    "Turkey": "Türkiye",
    "Cape Verde": "Cabo Verde",
    "Columbia": "Colombia",
    "Ivory Coast": "Côte d'Ivoire",
    "Iran": "IR Iran",
    "Curacao": "Curaçao",
}

MONTHS = {
    "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
    "July": 7, "August": 8, "September": 9, "October": 10, "November": 11,
    "December": 12,
}


def canon(name: str) -> str:
    name = (name or "").strip()
    return TEAM_ALIASES.get(name, name)


def split_owner(raw: str):
    """'Mexico (Sean)' -> ('Mexico', 'Sean')."""
    m = re.match(r"^(.*?)\s*\(([^)]*)\)\s*$", raw.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw.strip(), None


def parse_date(text: str) -> str | None:
    """'11th June' -> '2026-06-11' (the tournament runs in 2026)."""
    if not text:
        return None
    m = re.match(r"(\d{1,2})\w*\s+([A-Za-z]+)", text.strip())
    if not m:
        return None
    day = int(m.group(1))
    month = MONTHS.get(m.group(2).capitalize())
    if not month:
        return None
    return f"2026-{month:02d}-{day:02d}"


def fmt_time(value) -> str | None:
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    return value


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    teams_ws = wb["Teams"]
    tr = wb["Tracker"]

    # --- Paid status keyed by canonical team name (from the Teams sheet) ---
    paid_by_team: dict[str, bool] = {}
    for r in range(2, 50):
        team = teams_ws.cell(r, 2).value
        if not team:
            continue
        paid_by_team[canon(team)] = bool(teams_ws.cell(r, 3).value)

    # --- Canonical teams: group + owner, from the GroupStandings helper (BB:BL) ---
    teams = []
    team_group = {}
    for r in range(5, 53):
        group = tr.cell(r, 54).value  # BB
        raw = tr.cell(r, 55).value    # BC
        if not (group and raw):
            continue
        country, owner_short = split_owner(raw)
        team_group[country] = group
        teams.append({
            "country": country,
            "group": group,
            "owner_short": owner_short,
            "owner": OWNER_FULL_NAMES.get(owner_short, owner_short),
            "paid": paid_by_team.get(country, False),
        })
    teams.sort(key=lambda t: (t["group"], t["country"]))

    # --- Group fixtures (GroupFixtures table A4:L76) ---
    fixtures = []
    for r in range(5, 77):
        group = tr.cell(r, 1).value
        match = tr.cell(r, 2).value
        if not (group and match):
            continue
        fixtures.append({
            "match": str(match),
            "group": group,
            "home": canon(tr.cell(r, 3).value),
            "away": canon(tr.cell(r, 4).value),
            "date": parse_date(tr.cell(r, 11).value),
            "date_label": (tr.cell(r, 11).value or "").strip(),
            "ko": fmt_time(tr.cell(r, 12).value),
        })

    # --- Knockout dates/times keyed by match number (rows 92-122) ---
    ko_meta = {}
    for r in range(92, 123):
        no = tr.cell(r, 1).value
        if not isinstance(no, int):
            continue
        ko_meta[no] = {
            "round": tr.cell(r, 2).value,
            "date": parse_date(tr.cell(r, 10).value),
            "date_label": (tr.cell(r, 10).value or "").strip(),
            "ko": fmt_time(tr.cell(r, 11).value),
        }

    people = sorted({(t["owner"], t["owner_short"]) for t in teams})

    payload = {
        "tournament": "FIFA World Cup 2026",
        "groups": sorted({t["group"] for t in teams}),
        "teams": teams,
        "people": [{"name": n, "short": s} for n, s in people],
        "fixtures": fixtures,
        "knockout_meta": ko_meta,
        "prize_pot": {
            "total": 155,
            "stake_per_team": 5,
            "prizes": [
                {"label": "1st place", "amount": 50},
                {"label": "2nd place", "amount": 25},
                {"label": "MIND (charity)", "amount": 80},
            ],
        },
    }

    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {OUT.relative_to(ROOT)}")
    print(f"  teams={len(teams)} fixtures={len(fixtures)} people={len(people)} "
          f"ko_matches={len(ko_meta)}")


if __name__ == "__main__":
    main()
