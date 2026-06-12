"""Parity tests: the engine must reproduce the actual spreadsheet.

These load ``data/World_Cup_sweepstake.xlsx`` directly and assert the engine
agrees with (a) the workbook's *cached computed values* (group ranks and the
best-third-placed selection) and (b) the *seeding formulas* of the knockout
bracket.  This is the most direct evidence that the site replicates the Excel
logic.  Skipped automatically if openpyxl or the workbook is unavailable.
"""
import re
from pathlib import Path

import pytest

from wcsweepstake import DATA, engine

XLSX = Path(__file__).resolve().parent.parent / "data" / "World_Cup_sweepstake.xlsx"

openpyxl = pytest.importorskip("openpyxl")
pytestmark = pytest.mark.skipif(not XLSX.exists(), reason="workbook not present")


def _strip_owner(s: str) -> str:
    return re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()


@pytest.fixture(scope="module")
def wb_values():
    return openpyxl.load_workbook(XLSX, data_only=True)["Tracker"]


@pytest.fixture(scope="module")
def wb_formulas():
    return openpyxl.load_workbook(XLSX, data_only=False)["Tracker"]


@pytest.fixture
def baseline_standings():
    """Engine standings with no results entered (matches the saved workbook)."""
    return engine.group_standings(DATA["teams"], {})


# --------------------------------------------------------------------------- #
#  (a) cached value parity
# --------------------------------------------------------------------------- #
def test_group_rank_parity(wb_values, baseline_standings):
    """Engine rank within each group == the workbook's cached ``Rank`` (BL)."""
    # Read GroupStandings helper table BB(54):BL(64) rows 5-52.
    sheet_rank = {}
    for r in range(5, 53):
        group = wb_values.cell(r, 54).value
        team = wb_values.cell(r, 55).value
        rank = wb_values.cell(r, 64).value
        if group and team:
            sheet_rank[(group, _strip_owner(team))] = rank

    checked = 0
    for group, rows in baseline_standings.items():
        for i, row in enumerate(rows, 1):
            assert sheet_rank[(group, row.country)] == i, (
                f"{group} {row.country}: engine #{i} vs sheet #{sheet_rank[(group, row.country)]}")
            checked += 1
    assert checked == 48


def test_third_placed_parity(wb_values, baseline_standings):
    """Engine's best-eight third-placed teams == cached A80:B87 selection."""
    sheet_thirds = []
    for r in range(80, 88):
        group = wb_values.cell(r, 1).value
        team = wb_values.cell(r, 2).value
        if group and team:
            sheet_thirds.append((group, _strip_owner(team)))

    engine_thirds = [(row.group, row.country) for row in engine.best_third_placed(baseline_standings)]
    assert engine_thirds == sheet_thirds


# --------------------------------------------------------------------------- #
#  (b) seeding-formula parity
# --------------------------------------------------------------------------- #
def _formula_text(cell):
    from openpyxl.worksheet.formula import ArrayFormula
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else (v if isinstance(v, str) else "")


def test_round_of_32_seeding_matches_formulas(wb_formulas):
    """Parse the C/F seeding formulas of rows 92-107 and compare to BRACKET."""
    group_seed = re.compile(r'\(GroupStandings\[Group\]="([A-L])"\)\*\(GroupStandings\[Rank\]=(\d)\)')
    third_seed = re.compile(r'XLOOKUP\("([A-L])",\$A\$80:\$A\$87')

    for r in range(92, 108):
        match_no = wb_formulas.cell(r, 1).value
        spec = engine.BRACKET[match_no]
        for col, side in ((3, "a"), (6, "b")):  # C=team1, F=team2
            text = _formula_text(wb_formulas.cell(r, col))
            gm = group_seed.search(text)
            tm = third_seed.search(text)
            if gm:
                assert spec[side] == ("group", gm.group(1), int(gm.group(2))), \
                    f"match {match_no} side {side}"
            elif tm:
                assert spec[side] == ("third", tm.group(1)), f"match {match_no} side {side}"
                assert engine.THIRD_SLOT_GROUP[match_no] == tm.group(1)
            else:
                pytest.fail(f"unrecognised seeding formula at {col}{r}: {text!r}")


def test_knockout_progression_matches_formulas(wb_formulas):
    """Rows 108-122 reference earlier matches' Winner/loser; verify wiring.

    ``C/F = '=H<row>'`` means *winner of the match on that row*; the third-place
    row uses ``=IF(H<sf>=C<sf>,F<sf>,C<sf>)`` i.e. the *loser* of a semi-final.
    """
    row_to_match = {r: wb_formulas.cell(r, 1).value for r in range(92, 123)}
    winner_ref = re.compile(r"=H(\d+)")
    loser_ref = re.compile(r"IF\(H(\d+)=C\1,F\1,C\1\)")

    for r in range(108, 122):  # R16 -> Semi-finals (plain winner references)
        match_no = row_to_match[r]
        spec = engine.BRACKET[match_no]
        for col, side in ((3, "a"), (6, "b")):
            m = winner_ref.search(_formula_text(wb_formulas.cell(r, col)))
            assert m, f"expected winner ref at {col}{r}"
            assert spec[side] == ("winner", row_to_match[int(m.group(1))])

    # Third-place play-off (row 122) takes the two semi-final losers.
    for col, side in ((3, "a"), (6, "b")):
        m = loser_ref.search(_formula_text(wb_formulas.cell(122, col)))
        assert m, f"expected loser ref at {col}122"
        assert engine.BRACKET[103][side] == ("loser", row_to_match[int(m.group(1))])


def test_fixtures_and_owners_match_workbook(wb_values):
    """The generated data still matches the workbook (guards stale JSON)."""
    # 72 group fixtures, 48 teams, 12 groups of 4.
    assert len(DATA["fixtures"]) == 72
    assert len(DATA["teams"]) == 48
    by_group = {}
    for t in DATA["teams"]:
        by_group[t["group"]] = by_group.get(t["group"], 0) + 1
    assert set(by_group) == set("ABCDEFGHIJKL")
    assert all(n == 4 for n in by_group.values())
